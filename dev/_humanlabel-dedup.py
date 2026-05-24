"""
De-duplicate Approved HumanLabel values per Sub's rule:

  HumanLabel must be unique unless the DisplayName is also a duplicate.

For each duplicate group:
  1. Compute the voice-clean form of each row's DisplayName.
  2. If all voice-clean DisplayNames in the group match (e.g. CIG just appended
     "(Short Press)" / "(Long Press)" — which we strip as modifier tags),
     keep the duplicate HumanLabels — they describe the same chart bind.
  3. Otherwise (voice-DisplayNames differ): regenerate HumanLabel for each
     row whose voice-DisplayName is distinct, so each gets its own label.

Special-case `(Overflow)` — preserve it as a suffix on the HumanLabel,
since `_overflow` is a real distinct vjoy slot, not a modifier tag.
"""
import csv
import json
import re
import sys
from collections import defaultdict
from importlib.util import spec_from_file_location, module_from_spec
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

REPO = Path(r"E:\06. Dev Projects\Subs-Curated-Bindings")
CSV_IN = REPO / "tools" / "humanlabel-review.csv"
CSV_OUT = REPO / "tools" / "humanlabel-review.csv"  # in-place
XLSX_OUT = REPO / "tools" / "humanlabel-review.xlsx"

# Import voice helpers from the existing generator
spec = spec_from_file_location("vg", str(REPO / "tools" / "_humanlabel-voice-generate.py"))
vg = module_from_spec(spec); spec.loader.exec_module(vg)


def voice_clean_displayname(disp, xml_name):
    """Same as voice_transform but biased toward DisplayName as the source.
    Preserves '(Overflow)' suffix since it's a real distinct action variant."""
    if not disp or disp.startswith("ui_") or disp.startswith("@"):
        # Fall back to humanizing the XML name
        return vg.voice_transform(xml_name, disp)

    has_overflow = "(Overflow)" in disp
    clean = disp.replace(" (Overflow)", "").replace("(Overflow)", "").strip()

    voiced = vg.voice_transform(xml_name, clean)
    if has_overflow:
        voiced = f"{voiced} (Overflow)"
    return voiced


def main():
    with open(CSV_IN, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        cols = reader.fieldnames
        rows = list(reader)

    by_label = defaultdict(list)
    for i, r in enumerate(rows):
        by_label[r["Approved HumanLabel"]].append((i, r))

    fixes = []
    for label, group in by_label.items():
        if len(group) == 1:
            continue

        # Compute voice-clean DisplayName per row
        voiced = []
        for i, r in group:
            v = voice_clean_displayname(r["DisplayName"], r["XMLActionName"])
            is_chart = r["Source"].startswith("chart:")
            has_overflow = "(Overflow)" in r["DisplayName"]
            voiced.append((i, r, v, is_chart, has_overflow))

        # If all voice-DisplayNames match, allow the duplicate (same bind)
        distinct_voices = set(v for _, _, v, _, _ in voiced if v)
        if len(distinct_voices) <= 1:
            continue  # legit duplicate per Sub's rule

        # Find an anchor. Priority:
        #   1. chart-sourced + non-empty DisplayName + no overflow + label is NOT a generic single word
        #   2. chart-sourced + non-empty DisplayName + no overflow
        #   3. chart-sourced + non-empty DisplayName
        #   4. chart-sourced (any)
        # Generic-single-word labels (e.g. "Inc.", "Dec.", "Toggle") are bad
        # anchors because they're not distinctive enough to preserve.
        def is_generic_label(label):
            return len(label.split()) <= 1
        anchor = None
        for entry in voiced:
            if entry[3] and entry[1]["DisplayName"].strip() and not entry[4] and not is_generic_label(entry[1]["Approved HumanLabel"]):
                anchor = entry; break
        if anchor is None:
            for entry in voiced:
                if entry[3] and entry[1]["DisplayName"].strip() and not entry[4]:
                    anchor = entry; break
        if anchor is None:
            for entry in voiced:
                if entry[3] and entry[1]["DisplayName"].strip():
                    anchor = entry; break
        if anchor is None:
            for entry in voiced:
                if entry[3]:
                    anchor = entry; break

        anchor_label = anchor[1]["Approved HumanLabel"] if anchor else None

        for entry in voiced:
            i, r, v, is_chart, has_overflow = entry
            if anchor is not None and i == anchor[0]:
                continue  # leave anchor alone

            # Decide replacement
            if has_overflow and anchor_label and "(Overflow)" not in anchor_label:
                new_label = f"{anchor_label} (Overflow)"
            elif is_chart and v:
                # Non-anchor chart row — try voice-DisplayName first
                new_label = v
            elif v:
                new_label = v
            else:
                continue  # nothing to use

            if new_label != r["Approved HumanLabel"]:
                fixes.append((i, r["XMLActionName"], r["DisplayName"], r["Approved HumanLabel"], new_label))
                rows[i]["Approved HumanLabel"] = new_label

    print(f"Applied {len(fixes)} dedup fixes")
    for idx, (i, xml, disp, before, after) in enumerate(fixes[:40]):
        print(f"  row {i+2:3d} {xml:42s} disp={disp[:35]!r:38s}  '{before}' -> '{after}'")
    if len(fixes) > 40:
        print(f"  ... and {len(fixes) - 40} more")

    # Re-verify: are there still any remaining mismatched duplicates?
    by_label2 = defaultdict(list)
    for r in rows:
        by_label2[r["Approved HumanLabel"]].append(r)
    remaining = []
    for label, group in by_label2.items():
        if len(group) <= 1:
            continue
        voices = {voice_clean_displayname(g["DisplayName"], g["XMLActionName"]) for g in group}
        if len(voices) > 1:
            remaining.append((label, group, voices))
    if remaining:
        print(f"\nWARN: {len(remaining)} duplicate groups still have distinct voiced DisplayNames")
        for label, group, voices in remaining[:5]:
            print(f"  {label!r} ({len(group)} rows): voiced={voices}")

    # Save CSV
    with open(CSV_OUT, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        w.writerows(rows)
    print(f"\nWrote {CSV_OUT}")

    # Re-generate xlsx
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active; ws.title = "HumanLabel Review"
    ws.append(cols)
    for r in rows:
        ws.append([r[c] for c in cols])
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"
    edit_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    conf_fills = {
        "high": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "medium": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "low": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    }
    # column B (idx 2) = Approved, column F (idx 6) = Confidence
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=2).fill = edit_fill
        conf = ws.cell(row=r, column=6).value
        if conf in conf_fills:
            ws.cell(row=r, column=6).fill = conf_fills[conf]
    widths = {1: 36, 2: 32, 3: 42, 4: 22, 5: 32, 6: 11, 7: 38, 8: 26, 9: 26, 10: 26, 11: 18}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.auto_filter.ref = ws.dimensions
    wb.save(XLSX_OUT)
    print(f"Wrote {XLSX_OUT}")


if __name__ == "__main__":
    main()
